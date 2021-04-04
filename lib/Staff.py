#!/usr/bin/env python3
import csv
import os
import time
import json
import string
import random
import base64
from datetime import date
from email.mime.text import MIMEText
import hashlib

import openpyxl
import requests
import urllib.request
import urllib.parse as urlparse
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from datetime import datetime
from Authentication import Authentication

class Staff:
    '''
    A class for performing user related tasks. A path to an excel file containing
    a list of users is required (usersFilePath), as well as an instance of
    the Authentication class to generate the necessary keys (auth)

    VARIABLES
        users: A Python list containing employees as Python dictionaries
        auth: An Authentication object passed during Employee instance creation
    '''

    ##### VARIABLES #####
    users = []
    auth = {}
    log = open('../log/staff.log', 'w+')
    

    ##### FUNCTIONS #####
    def __init__(self, usersFilePath, auth):
        #self.getUserList(usersFilePath)
        self.auth = auth


    def getUserList(self, usersFilePath): 
        if os.path.isfile(usersFilePath): self.useUserList(usersFilePath)
        else:
            path = input(f'\nNo user file found.  Enter path to user file or press ENTER to input users manually: ') 
            if path: self.useUserList(path)
            else: self.useManualInput()


    def useUserList(self, usersFilePath):
        ''' 
        Retrieves user data from excel file
        
        Args
            usersFilePath: File path to user data   
        Returns 
            Formatted list of users as dictionaries
        '''

        # Open excel file
        # Split col A into first and last name
        # col T (Personal Email)
        # col I (Hiring Manager)
        # col L (Department)
        # col M (Title)
        # col G ( ) Temp or FT-Admin *i
        print('getting users...')
        excelFile = openpyxl.load_workbook(usersFilePath)
        sheet = excelFile['Pending Start']
        print('excel sheet opened...')
        # Get today's date
        today = date.today()
        startDate = date.today()
        
        for cell in sheet['B']:
            
            row = cell.row          
            # If column B cell is not empty, get start date
            if cell.value:               
                try:
                    startDate = str(cell.value)
                    year = int(startDate[0:4])
                    month = int(startDate[5:7])
                    day = int(startDate[8:10])
                    startDate = date(year, month, day)
                except:
                    # TODO Add to log file
                    self.log.write('getNewHires ==> INCORRECT DATE FORMAT! Row ' + str(cell.row) + '\n')
                    startDate = date.today()
            
            # Add all newHires starting in San Francisco office to user list
            startLocation = sheet.cell(row=row, column=11).value
            if startDate > today and (startLocation == 'San Francisco' or startLocation == 'Obscura'):
                name = sheet.cell(row=row, column=1).value.split(' ')
                personalEmail = sheet.cell(row=row, column=20).value
                department = sheet.cell(row=row, column=12).value
                title = sheet.cell(row=row, column=13).value
                manager = sheet.cell(row=row, column=9).value
                status = sheet.cell(row=row, column=7).value
                self.users.append({
                    'firstName': name[0],
                    'lastName': name[1],
                    'personalEmail': personalEmail,
                    'workEmail': f'{name[0].lower()}.{name[1].lower()}@obscuradigital.com',
                    'department': department,
                    'title': title,
                    'manager': manager,
                    'status': status.lower(),
                    'startDate': startDate,
                })
        self.log.write('\n')
        self.log.write('getNewHires ==> Found the following users in Onboarding Tracker:\n\n')
        for user in self.users: self.log.write(str(user) + '\n\n')   


    def useManualInput(self):
        ''' 
        Retrieves user data manually from administrator
        '''
        done = False
        while not done:

            ## Enter user info
            user = {}
            print('\nEnter new user information')
            user['firstName'] = input('First name: ')
            user['lastName'] = input('Last name: ')
            user['personalEmail'] = input('Personal email: ')
            user['workEmail'] = input('Work email: ')
            user['department'] = input('Department: ')
            user['title'] = input('Job title: ')
            user['manager'] = input('Manager: ')
            user['status'] = input('Status (ft or temp): ')
            user['startDate'] = input('Start date (mm/dd/yyyy):')
            
            ## Validate entry
            print('\nYou have entered the following new user.')
            print(user)
            confirm = input('Is this correct? (Y or N): ')
            if confirm.lower() == 'n': continue

            ## Add new user to users list
            self.users.append(user)
            confirm = input('Would you like to add another user? (Y or N): ')
            if confirm.lower() == 'n': 
                done = True


    def getAllGSuiteUsers(self, nextPageToken=0): 
        ''' 
        Retrieves all users from the domain

        Returns
            Object containing all users from obscuradigital.com
            False if error during operation
        '''
        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()
        # Make API request to add new user

        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }

        response = requests.get('https://www.googleapis.com/admin/directory/v1/users?domain=detectnow.com', headers=HEADERS)
        allUsers = response.json()
        final = []
        for user in allUsers['users']:
            self.log.write(user['primaryEmail'] + "\n")
            final.append(user['primaryEmail'])
        while "nextPageToken" in allUsers: 
            response = requests.get('https://www.googleapis.com/admin/directory/v1/users?domain=detectnow.com', headers=HEADERS, params={ 'pageToken': allUsers['nextPageToken'] })
            allUsers = response.json()
            for user in allUsers['users']:
                self.log.write(user['primaryEmail'] + "\n")
                final.append(user['primaryEmail'])
        return final

    def updateGSuiteEmail(self, oldEmail, newEmail):
        '''
        Updates user email in GSuite
        '''
        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()   
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }
        BODY = { "primaryEmail": newEmail }

        # Make API request to change primary email
        response = requests.put(f'https://www.googleapis.com/admin/directory/v1/users/{oldEmail}', headers=HEADERS, json=BODY)
        print(response.json())
        # time.sleep(1)

        # Now let's add the old email as an alias
        # BODY = { "alias": oldEmail }
        # response = requests.post(f'https://www.googleapis.com/admin/directory/v1/users/{newEmail}/aliases', headers=HEADERS, json=BODY)


    def addGSuiteUser(self, user): 
        ''' 
        Generates an email address in gSuite for new hire
        
        Arguments
            user: Dictionary of user data to be added

        Returns
            Object containing created user from Google
            False if error during operation
        '''
        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()
        # Make API request to add new user
        USER_DATA = {
            "name": {
                "familyName": user['lastName'],
                "givenName": user['firstName']
            },
            "password": self.auth.keys['default_password'],
            "primaryEmail": user['workEmail'],
        }

        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }
        addedUser = requests.post('https://www.googleapis.com/admin/directory/v1/users', headers=HEADERS, json=USER_DATA)
        if addedUser.status_code == 200:
            self.log.write(f"addGSuiteUser ==> Email creation successful for {user['firstName']} {user['lastName']}\n{addedUser.json()}\n")
            return addedUser.json()
        else:
            self.log.write(f"addGSuiteUser ==>  Something went wrong with the request. {user['firstName']} {user['lastName']} not added\n{addedUser.json()}\n")
            return False


    def removeGSuiteUser(self, userEmail):
        ''' 
        Removes a user from gSuite
        
        Arguments
            userEmail: Email address of the user to be removed

        Returns
            True if successfully removed
            False if error during operation
        '''

        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()

        # Make API request to remove user
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }
        removedUser = requests.delete(f'https://www.googleapis.com/admin/directory/v1/users/{userEmail}', headers=HEADERS)
        if removedUser.status_code == 204:
            self.log.write(f"removeGSuiteUser ==> Removed {userEmail} from Obscura domain\n")
            return True
        else:
            self.log.write(f"removeGSuiteUser ==>  Something went wrong with the request. {userEmail} not removed\n{removedUser}\n")
            return False

 
    def generatePassword(self):
        l = string.ascii_lowercase + string.ascii_uppercase
        d = string.digits
        newPassword = ''.join(random.sample(l, 10)) + ''.join(random.sample(d, 5))
        return newPassword

        
    def resetGSuitePassword(self, user):
        ''' 
        Generates a random password and sends an email to the administrator with credentials
        
        Arguments
            user: Dictionary of user data

        Returns
            updatedUser if successfully reset
            False if error during operation
        '''
        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()
        newPassword = self.generatePassword()
        print(newPassword)
        print(user['primaryEmail'])
        # Define variables
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }
        USER_DATA = {
            "password": newPassword,
            "changePasswordAtNextLogin": 'true',
        }
        
        # Make API request to reset password
        updatedUser = requests.put(f'https://www.googleapis.com/admin/directory/v1/users/{user["primaryEmail"]}', headers=HEADERS, json=USER_DATA)
        
        csv = open('../files/email_passwords.csv', 'a+')
        if updatedUser.status_code == 200:
            msg = (
                f'Greetings from your IT department,'
                f'As a reminder, during the next step in our migration to MSG Platforms, we’ll be separating the '
                f'connection between MSG\'s active directory and G Suite, and as part of this separation, you '
                f'will be required to use a new password to access G Suite starting tomorrow at 6:00pm. We’ve provided you with your new password below: '
                f'\n{newPassword}'
                f'\nPlease reach out to the MSG Service Desk at 212-465-6565 if you experience any issues. Thank you for your cooperation, and have a great day!'
            )
            #self.log.write(f"resetPassword ==> Reset password for {user['primaryEmail']} from Obscura domain\n")
            csv.write(f'{user["primaryEmail"]},{newPassword}\n')
            csv.close()
            self.sendEmail(f"{self.auth.keys['admin_email']}, {user['primaryEmail']}", f"Password reset for {user['primaryEmail']}", msg, 'plain', user)
            print(updatedUser)
            return updatedUser
        else:
            #self.log.write(f"resetPassword ==>  Something went wrong with the request. {user['workEmail']} password was not changed.\n")
            print(updatedUser)
            return updatedUser


    def addToGSuiteGroups(self, user, groups): 
        '''
        Adds a user to email group aliases

        Arguments
            user: The person who will be added to groups
            groups: A list of the group names to which the user will be added
        Returns
            Google response in json object, if successfully added
            False if error occurs
        '''

        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()
        
        # Define variables
        USER_DATA = {
            "email": user['workEmail'],
            "role": 'MEMBER',
        }
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }

        # Make an API request for each group in the list
        for group in groups:
            addedUser = requests.post(f'https://www.googleapis.com/admin/directory/v1/groups/{group}@obscuradigital.com/members', headers=HEADERS, json=USER_DATA)
            if addedUser.status_code == 200:
                self.log.write(f"addToGSuiteGroups ==> {user['firstName']} {user['lastName']} successfully added to {group}@obscuradigital.com\n{addedUser.json()}\n")
                return addedUser.json()
            else:
                self.log.write(f"addToGSuiteGroups ==>  Something went wrong with the request. {user['firstName']} {user['lastName']} not added to {group}@obscuradigital.com\n{addedUser.json()}\n")
                return False


    def createSlackAccount(self, user):
        '''
        Generates an email address in gSuite for new hire

        Arguments
            user: User to send Slack invitation
        Returns
            True if invitation sent successfully
            False if error occured
        '''

        # Define variables
        URL = 'https://slack.com/api/users.admin.invite'
        DATA = {    
            'email': user['workEmail'],
            'channels':'C9FUXFA3A',
            'real_name': f"{user['firstName']} {user['lastName']}",  
        }
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['slack']['ACCESS_TOKEN'],
            'Content-type': 'application/x-www-form-urlencoded',
        }

        # Make POST request
        response = requests.post(url = URL, data = DATA, headers = HEADERS)
        res = response.json()
        
        # Print result
        if res['ok'] == True:
            self.log.write('createSlackAccount ==> Slack invitation successful for ' + DATA['real_name'] + '\n')
            return True
        else: 
            self.log.write(f"createSlackAccount ==> Something went wrong. Slack invite not sent to {user['firstName']} {user['lastName']}.\n{res['error']}\n")
            return False

    def getAllSlackUsers(self):
        '''
        Returns a list of all Slack users
        '''

        # Define variables
        URL = 'https://slack.com/api/users.list'
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['slack']['ACCESS_TOKEN'],
            'Content-type': 'application/x-www-form-urlencoded',
        }

        # Make GET request
        response = requests.get(url = URL, headers = HEADERS)
        res = response.json()

        # Print result
        if res['ok'] == True:
            for user in res['members']:
                if user['deleted'] == True or user['is_bot'] == True or user['name'] == 'slackbot':
                    res['members'].remove(user)
            print(f'There are {len(res["members"])} active users')
            self.log.write(f'getAllSlackUsers ==> {res}\n')
            return res['members']
        else: 
            self.log.write(f"getAllSlackUsers ==> {res}\n")
            return False


    def updateSlackEmail(self, user):
        '''
        Updates email address of Slack user from obscuradigital.com to msg.com
        '''
        
        # Define variables
        updatedEmail = user['profile']['email'].replace('detectnow', 'detect')
        URL = 'https://slack.com/api/users.profile.set'
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['slack']['ACCESS_TOKEN'],
            'Content-type': 'application/json; charset=utf-8',
            'X-Slack-User': user['id'],
        }
        DATA = {
            "profile": { "email": updatedEmail }
        }
        
        # Make POST request
        response = requests.post(url = URL, headers = HEADERS, json = DATA)
        res = response.json()

        # Print result
        if res['ok'] == True:
            print(f"{user['profile']['email']} updated to {updatedEmail}")
            self.log.write(f'updateSlackEmail ==> {res}\n')
            return res
        else: 
            print(f"{user['profile']['email']} not updated")
            self.log.write(f"updateSlackEmail ==> {res}\n")
            return res



    def createZoomAccount(self, user): 
        '''
        Creates an account in Zoom for user
            https://marketplace.zoom.us/docs/api-reference/zoom-api/users/usercreate

        Arguments
            user: User of the new Zoom account
        Returns
            Response from Zoom as json object, if account created successfully
            False if error occurs
        '''

         # Define variables
        URL = 'https://api.zoom.us/v2/users'
        DATA = {    
            'action': 'create',
            'user_info': {
                'email': user['workEmail'],
                'type': '2',
                'first_name': user['firstName'],
                'last_name': user['lastName']
            } 
        }
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['zoom']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }

        # Make POST request
        addedUser = requests.post(url = URL, json = DATA, headers = HEADERS)
        
        # Print result
        if addedUser.status_code == 201:
            self.log.write(f"createZoomAccount ==> Pro Zoom account successfully created for {user['firstName']} {user['lastName']}\n{addedUser.json()}\n")
            return addedUser.json()
   
        elif addedUser.status_code == 200:
            DATA['user_info']['type'] = '1'
            addedUser = requests.post(url = URL, json = DATA, headers = HEADERS)
            if addedUser.status_code == 201:
                self.log.write(f"createZoomAccount ==> Basic Zoom account successfully created for {user['firstName']} {user['lastName']}\n{addedUser.json()}\n")
                return addedUser.json()
            else: 
                self.log.write(f"createZoomAccount ==> ERROR {user['firstName']} {user['lastName']} {addedUser.json()}\n")
                return False
        else: 
            self.log.write(f"createZoomAccount ==> ERROR {user['firstName']} {user['lastName']} {addedUser.json()}\n")
            return False
        

    def createHarvestAccount(self, user): 
        '''
        Creates an account in Harvest for user

        Arguments
            user: User of the new Harvest account
        Returns
            Harvest response as json object if created successfully
            False if error occurred
        '''

        # Define variables
        URL = 'https://api.harvestapp.com/v2/users'
        DATA = {    
            'first_name': user['firstName'],
            'last_name': user['lastName'],
            'email': user['workEmail'],
            'roles': [user['title']],
        }
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['harvest']['ACCESS_TOKEN'],
            'Harvest-Account-Id': self.auth.keys['harvest']['ACCOUNT_ID'],
            'User-Agent': f"MyApp ({self.auth.keys['admin_email']})",
            'Content-type': 'application/json',
        }

        # Make POST request
        addedUser = requests.post(url = URL, json = DATA, headers = HEADERS)
        
        # Print result
        if addedUser.status_code == 201:
            self.log.write(f"createHarvestAccount ==> Harvest account successfully created for {user['firstName']} {user['lastName']}\n{addedUser.json()}\n")
            return addedUser.json()
        else:
            self.log.write(f"createHarvestAccount ==> Something went wrong with the request. Harvest user {user['firstName']} {user['lastName']} not created\n{addedUser.json()}\n")
            return False


    def sendEmail(self, destination, subject, msg, type='plain', user={}): 
        '''
        Sends an email to using Gmail messages API. Sender is the user with the access token
            https://developers.google.com/gmail/api/v1/reference/users/messages/send
        Arguments
            destination: The recipient's email address
            subject: The subject line of the email
            msg: The body of the email message
            type: Format of message, i.e. text, html, etc
            user: User object
        Returns
            True if email sent successfully
            False if not sent
        '''

        if not self.auth.keys['google']['ACCESS_TOKEN']: self.auth.getGoogleAccessToken()
        
        # Define variables
        URL = 'https://www.googleapis.com/gmail/v1/users/me/messages/send'
        HEADERS = {
            'Authorization': 'Bearer ' + self.auth.keys['google']['ACCESS_TOKEN'],
            'Content-type': 'application/json',
        }
        if msg == '': msg = self.createWelcomeEmail(user)
        message = MIMEText(msg, type)
        message['to'] = destination
        message['from'] = self.auth.keys['admin_email']
        message['subject'] = subject
        DATA = {'raw': base64.urlsafe_b64encode( message.as_string().encode('UTF-8') ).decode('ascii')}
       
        # Make POST request
        response = requests.post(url = URL, json = DATA, headers = HEADERS)
        res = response.json()
              
        # Print result
        if response.status_code == 200:
            print(f"sendEmail ==> Email sent to {destination}\n{res}\n\n")
            return True
        else: 
            print(f"sendEmail ==> ERROR {res}\n\n")
            return False

    
    def createWelcomeEmail(self, user):
        email = (
            '<html><body>'
            '<h2>A welcome from your IT team!</h2>'
            f'<p>Hello, {user["firstName"]}. We are thrilled that you will be joining us soon. '
            'Here are a few things that will help you be prepared for your first day:</p>'
            
            '<h3>Email</h3>'
            f'<p>Your new Obscura email account has been created. Your address is {user["workEmail"]} and the default password is {self.auth.keys["default_password"]}. '
            'You can login at <a href="https://mail.google.com">mail.google.com</a> to start setting up your inbox</p>'
            
            '<h3>Slack</h3>'
            '<p>Your inbox should contain an invitation to our <a href="https://obscurahq.slack.com">Slack channel</a>. '
            'The "all sf" channel is used primarily for work related announcements, whereas the "watercooler" channel '
            'is the space for more personal conversations.  Feel free to drop a line any time!</p>'

            '<h3>Zoom</h3>'
            '<p>Your inbox should also contain an invitation to your new Zoom account. We use Zoom a lot to communicate with vendors and the MSG offices in New York. '
            'Once you get your computer, you should download the <a href="https://zoom.us/download">Zoom Client for Meetings</a> app for easy access to Zoom\'s many features. '
            'If you use Chrome, <a href="https://chrome.google.com/webstore/detail/zoom-scheduler/kgjfgplpablkjnlkjmjdecgdpfankdle?hl=en-US">click here</a> for a nifty Zoom extension that is great for creating meetings in Google calendar.</p>'
        )

        # if user['status'] != 'temp': email += (
        #     '<h3>Harvest</h3>'
        #     '<p>Harvest is the software we use for time tracking. That will be the last '
        #     'invitation you\'ll find in your inbox.  <a href="https://youtu.be/AA1oUDPO6Ns">Watch this video</a> to get up to speed on how '
        #     'Obscura uses this platform.</p></br>'
        # )

        email += (
            f'<p>New Hire orientation will be at 10:00am on {user["startDate"].strftime("%B %d, %Y")}.  We look forward to meeting you then.  If you have any questions about this email or need '
            'assistance with setting up your accounts, you can contact us by sending an email to ithelp@obscuradigital.com.</p></br>'
            '<p>Welcome aboard!</p></br></br>'
            '<p>Your IT team</p>'
            '</body></html>'
        )
        return email

    
def employeeTest():
    # Define file path to new hire data
    pathToUsers = ''

    # Get admin keys - REQUIRED to run this script!
    #auth = Authentication('../../_keys/adminKeys.py')
    auth = {}
    # Format the list as an Employee object
    staff = Staff(pathToUsers, auth)

    # # For each new hire on the list, do the following:  
    # # ## Comment out unnecessary tasks ## #
    for newHire in staff.users:
    #     print( staff.addGSuiteUser(newHire) )
    #     staff.addToGSuiteGroups(newHire, ['obscura', 'od-temps'])
    #     print( staff.createSlackAccount(newHire) )
    #     print( staff.createZoomAccount(newHire) )
    #     print( staff.createHarvestAccount(newHire) )
        print( staff.sendEmail(staff.auth.keys['admin_email'], f"Welcome to Obscura Digital, {newHire['firstName']}", '', 'html', newHire) )
    #     print( staff.removeGSuiteUser(newHire['workEmail']) )
    # print( staff.sendEmail(staff.auth.keys['admin_email'], 'New hire(s) added', 'TEST COMPLETED', 'text') ) # Confirmation email to newhires group

    # # COMMENT OUT TO STOP REMOVAL OF ADMIN KEYS AFTER SCRIPT EXIT (not recommended)
    # if staff.auth.removeKeys():
    #     print('Keys safely removed from working directory\n')
    # else: print('Keys not removed from working directory\n')
    # return None

if __name__ == '__main__':
    employeeTest()     

